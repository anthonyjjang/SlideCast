import openpyxl
import subprocess
import os

EXCEL_FILE = "발표_스크립트_시트.xlsx"
SHEET_NAME = "발표 스크립트"
OUTPUT_TXT = "발표_스크립트_전체.txt"
OUTPUT_AUDIO = "발표_음성_Reed.aiff"
VOICE = "Reed"  # 한국어 음성: Yuna, Eddy, Flo, Reed, Sandy, Shelley, Rocko 중 선택

# 1. 엑셀에서 G열 스크립트 추출
print("엑셀 파일 읽는 중...")
wb = openpyxl.load_workbook(EXCEL_FILE)
ws = wb[SHEET_NAME]

texts = []
for row in range(2, ws.max_row + 1):
    val = ws.cell(row=row, column=7).value
    if val and str(val).strip():
        texts.append(str(val).strip())

full_text = "\n\n".join(texts)
print(f"  총 {len(texts)}개 슬라이드, {len(full_text):,}자 추출 완료")

# 2. 텍스트 파일로 저장
with open(OUTPUT_TXT, "w", encoding="utf-8") as f:
    f.write(full_text)
print(f"  텍스트 저장: {OUTPUT_TXT}")

# 3. say 명령어로 음성 생성
print(f"\n음성 생성 중... (음성: {VOICE})")
print("약 40-50분 분량이라 시간이 걸릴 수 있습니다. 잠시 기다려주세요...")

result = subprocess.run(
    ["say", "-v", VOICE, "-f", OUTPUT_TXT, "-o", OUTPUT_AUDIO],
    capture_output=True,
    text=True
)

if result.returncode == 0:
    size_mb = os.path.getsize(OUTPUT_AUDIO) / (1024 * 1024)
    print(f"\n완료! 음성 파일 생성됨: {OUTPUT_AUDIO} ({size_mb:.1f} MB)")
    print("\n재생하려면:")
    print(f"  open {OUTPUT_AUDIO}")
    print("\n다른 음성으로 바꾸려면 VOICE 변수를 수정하세요:")
    print("  여성: Yuna, Flo, Sandy, Shelley")
    print("  남성: Eddy, Reed, Rocko, Grandpa")
else:
    print(f"오류 발생: {result.stderr}")
